from openpyxl import *
from openpyxl.styles import *
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import *
import pandas as pd
from datetime import date
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment


def merge(master_sheet_filepath, formatted_dpp_filepath):
    print(f'Master filepath: {master_sheet_filepath}')
    print(f'Formatted DPP filepath: {formatted_dpp_filepath}')

    master_wb = load_workbook(master_sheet_filepath)
    master_ws = master_wb.active
    dpp_wb = load_workbook(formatted_dpp_filepath)
    dpp_ws = dpp_wb.active

    # Get unique dispatch numbers
    m_dispatch = set(cell.value for cell in master_ws['C'] if cell.value is not None)
    d_dispatch = set(cell.value for cell in dpp_ws['C'] if cell.value is not None)
    unique_to_dpp = d_dispatch - m_dispatch
    print(f'Unique to DPP: {unique_to_dpp}')

    # Ignore columns that will contain manual entries
    ignore_cols = [
        "Missing Information", "Zone Uplift", 
        "Overnight", "Shift Uplift", "PM Confirmation", 
        "Corrected SKU", "Completed Date",
        "Scheduled Date", "Admin Status", "Notes"
    ]
    ignore_cols_indices = [master_ws[1].index(cell) for cell in master_ws[1] if cell.value in ignore_cols]

    # Ignore columns with default values set after merge
    ignore_cols_default_values = [
        'Admin Status'
    ]
    ignore_cols_default_values_indices = [master_ws[1].index(cell) for cell in master_ws[1] if cell.value in ignore_cols_default_values]

    # Find the Missing Information column index
    missing_info_col_idx = None
    for idx, cell in enumerate(master_ws[1]): 
        if cell.value == "Missing Information":
            missing_info_col_idx = idx
            print(f'Missing info index: {missing_info_col_idx}')
            break

    if missing_info_col_idx is None:
        print("Missing Information column not found in the header row.")
    

    # Create data validation object (dropdown) and Admin Status Dropdown options
    admin_status_dropdown_list = [
        'New', 'Acknowledged', 
        'Scheduled', 'On Site', 
        'Stuck', 'Postponed', 
        'Incomplete', 'Complete'
    ]
    admin_status_dropdown = DataValidation(type='list', formula1='"{}"'.format(','.join(admin_status_dropdown_list)))


    # Append new rows to the master sheet
    # Apply styling and formatting to new rows
    new_rows = []
    if unique_to_dpp:
        for row in dpp_ws.iter_rows(min_row=2):
            if row[2].value in unique_to_dpp:
                row_values = [cell.value for cell in row]
                master_ws.append(row_values)
                new_row = master_ws[master_ws.max_row]
                style_row(new_row, ignore_cols_indices, ignore_cols_default_values_indices, missing_info_col_idx)
                new_rows.append(new_row)

    # Apply formatting only to new rows
    if new_rows:
        format_cells(new_rows)


    # Add dropdown to the master sheet
    create_admin_status_dropdown(master_ws,admin_status_dropdown)
    master_wb.save(master_sheet_filepath)
    print('Merging, styling, and formatting of new rows complete.')


# Style new rows
def style_row(row, ignore_cols_indices, ignore_cols_default_values_indices, missing_info_col_idx):
    green_fill = PatternFill(start_color="FF90EE90", end_color="FF90EE90", fill_type="solid")
    red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
    grey_fill = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")
    has_missing = False

    for idx, cell in enumerate(row):
        if idx in ignore_cols_default_values_indices:
            cell.fill = green_fill
            continue
        if idx in ignore_cols_indices:
            cell.fill = grey_fill
            continue
        cell.fill = green_fill
        if cell.value is None:
            cell.fill = red_fill
            has_missing = True

    if has_missing:
        row[missing_info_col_idx].value = "Yes"
        row[missing_info_col_idx].fill = red_fill


# Format new cells
def format_cells(rows):
    # Apply text wrap and auto-fit column widths only to specified rows
    for row in rows:
        for cell in row:
            cell.alignment = Alignment(wrapText=True)
            max_length = len(str(cell.value)) if cell.value else 0
            adjusted_width = (max_length + 2) * 1.1
            cell.parent.column_dimensions[cell.column_letter].width = max(adjusted_width, cell.parent.column_dimensions[cell.column_letter].width)


def create_admin_status_dropdown(master_ws, dropdown):
    column_letter = 'A'
    start_row = 2
    end_row = master_ws.max_row

    cell_range = f'{column_letter}{start_row}:{column_letter}{end_row}'
    for row in range(start_row, end_row + 1):
        cell = master_ws[f'{column_letter}{row}']
        if cell.value is None:
            cell.value = "New"

    master_ws.add_data_validation(dropdown)
    dropdown.add(cell_range)
